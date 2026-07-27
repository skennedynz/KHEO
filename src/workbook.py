"""
KHEO Version 1.0.0
workbook.py
"""

from pathlib import Path

from openpyxl import Workbook

from openpyxl.utils import get_column_letter

from src.styles import (
    title,
    header,
    value,
    setup_sheet,
)
from src.solar import (
    load_roof_geometry,
    total_array_capacity,
)

from src.energy import (
    annual_summary,
    monthly_energy_balance,
)

OUTPUT_FOLDER = Path("output")
OUTPUT_FILE = OUTPUT_FOLDER / "KHEO.xlsx"

def autofit_columns(worksheet):
    """
    Automatically adjust the width of each column to fit its contents.
    """

    for column in worksheet.columns:

        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:

            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass

        worksheet.column_dimensions[column_letter].width = max_length + 2

def build_workbook():

    OUTPUT_FOLDER.mkdir(exist_ok=True)

    wb = Workbook()

    # --------------------------------------------------
    # Worksheets
    # --------------------------------------------------

    home = wb.active
    home.title = "Home"

    inputs = wb.create_sheet("Inputs")
    energy = wb.create_sheet("Energy")
    solar = wb.create_sheet("Solar")
    hotwater = wb.create_sheet("Hot Water")
    battery = wb.create_sheet("Battery")
    financial = wb.create_sheet("Financial")
    dashboard = wb.create_sheet("Dashboard")

    worksheets = [
        home,
        inputs,
        energy,
        solar,
        hotwater,
        battery,
        financial,
        dashboard,
    ]

    for ws in worksheets:
        setup_sheet(ws)

    # --------------------------------------------------
    # ENERGY DATA
    # --------------------------------------------------

    summary = annual_summary()
    monthly = monthly_energy_balance()

    # --------------------------------------------------
    # HOME
    # --------------------------------------------------

    home["A1"] = "Kennedy Home Energy Optimiser"
    title(home["A1"])

    home["A3"] = "Property"
    header(home["A3"])

    home["A4"] = "Owner"
    home["B4"] = "Steve Kennedy"

    home["A5"] = "Location"
    home["B5"] = "Christchurch"

    home["A6"] = "House Area"
    home["B6"] = "247 m²"

    home["A7"] = "Occupants"
    home["B7"] = "2"

    for row in range(4, 8):
        value(home[f"A{row}"])
        value(home[f"B{row}"])

    # --------------------------------------------------

    home["A10"] = "Solar System"
    header(home["A10"])

    home["A11"] = "Panels"
    home["B11"] = "23 × 470 W"

    home["A12"] = "Array Size"
    home["B12"] = "10.81 kW"

    home["A13"] = "Orientation"
    home["B13"] = "14 NE / 9 NW"

    home["A14"] = "Inverter"
    home["B14"] = "Sigen 12 kW TP2"

    home["A15"] = "Battery"
    home["B15"] = "None Installed"

    home["A16"] = "EV Charger"
    home["B16"] = "Sigen 11 kW AC"

    for row in range(11, 17):
        value(home[f"A{row}"])
        value(home[f"B{row}"])

    # --------------------------------------------------

    home["A19"] = "Hot Water"
    header(home["A19"])

    home["A20"] = "Cylinder"
    home["B20"] = "Peter Cocks 300 L"

    home["A21"] = "Elements"
    home["B21"] = "2 × 3 kW"

    home["A22"] = "Thermostat"
    home["B22"] = "70 °C"

    home["A23"] = "Controllers"
    home["B23"] = "2 × Shelly Pro 1PM"

    for row in range(20, 24):
        value(home[f"A{row}"])
        value(home[f"B{row}"])

    # --------------------------------------------------

    home["D3"] = "Annual Summary"
    header(home["D3"])

    home["D4"] = "Annual Usage"
    home["E4"] = summary["usage"]

    home["D5"] = "Estimated Solar"
    home["E5"] = summary["solar"]

    home["D6"] = "Annual Balance"
    home["E6"] = summary["balance"]

    home["D7"] = "Solar Coverage"
    home["E7"] = summary["coverage"]
    home["E7"].number_format = "0.0%"

    for row in range(4, 8):
        value(home[f"D{row}"])
        value(home[f"E{row}"])

    # --------------------------------------------------
    # INPUTS
    # --------------------------------------------------

    inputs["A1"] = "Inputs"
    title(inputs["A1"])

    inputs["A3"] = "Parameter"
    inputs["B3"] = "Value"

    header(inputs["A3"])
    header(inputs["B3"])

    input_rows = [
        ("House Area (m²)", 247),
        ("Occupants", 2),
        ("Location", "Christchurch"),
        ("Panels", 23),
        ("Panel Rating (W)", 470),
        ("Array Size (kW)", 10.81),
        ("Orientation", "14 NE / 9 NW"),
        ("Inverter", "Sigen 12 kW TP2"),
        ("Battery", "None"),
        ("EV Charger", "Sigen 11 kW AC"),
        ("Hot Water Cylinder", "Peter Cocks 300 L"),
        ("Elements", "2 × 3 kW"),
        ("Thermostat (°C)", 70),
        ("Tariff", "Contact Good Weekends"),
    ]

    row = 4

    for name, val in input_rows:
        inputs[f"A{row}"] = name
        inputs[f"B{row}"] = val

        value(inputs[f"A{row}"])
        value(inputs[f"B{row}"])

        row += 1

    # --------------------------------------------------
    # ENERGY
    # --------------------------------------------------

    energy["A1"] = "Monthly Energy Balance"
    title(energy["A1"])

    headings = [
        "Month",
        "Usage (kWh)",
        "Solar (kWh)",
        "Direct Use (kWh)",
        "Grid Import (kWh)",
        "Grid Export (kWh)",
        "Balance (kWh)",
    ]

    col = 1

    for text in headings:

        cell = energy.cell(row=3, column=col)
        cell.value = text
        header(cell)

        col += 1

    row = 4

    for item in monthly:

        energy.cell(row=row, column=1).value = item["month"]
        energy.cell(row=row, column=2).value = item["usage"]
        energy.cell(row=row, column=3).value = item["solar"]
        energy.cell(row=row, column=4).value = item["direct_use"]
        energy.cell(row=row, column=5).value = item["grid_import"]
        energy.cell(row=row, column=6).value = item["grid_export"]
        energy.cell(row=row, column=7).value = item["balance"]

        for col in range(1, 8):
            value(energy.cell(row=row, column=col))

        row += 1

    energy[f"A{row}"] = "Annual Total"
    header(energy[f"A{row}"])

    energy[f"B{row}"] = summary["usage"]
    energy[f"C{row}"] = summary["solar"]

    energy[f"D{row}"] = sum(item["direct_use"] for item in monthly)
    energy[f"E{row}"] = sum(item["grid_import"] for item in monthly)
    energy[f"F{row}"] = sum(item["grid_export"] for item in monthly)

    energy[f"G{row}"] = summary["balance"]

    for col in range(2, 8):
        value(energy.cell(row=row, column=col))

    # --------------------------------------------------
    # SOLAR
    # --------------------------------------------------

    solar["A1"] = "Solar PV System"
    title(solar["A1"])

    headings = [
        "Roof",
        "Panels",
        "Panel Rating (kW)",
        "Array Size (kW)",
        "Tilt (°)",
        "Azimuth (°)",
    ]

    for col, text in enumerate(headings, start=1):

        cell = solar.cell(row=3, column=col)
        cell.value = text
        header(cell)

    roofs = load_roof_geometry()

    row = 4

    total_panels = 0
    total_kw = 0.0

    for roof in roofs:

        array_kw = roof["panels"] * roof["panel_rating"]

        solar.cell(row=row, column=1).value = roof["roof"]
        solar.cell(row=row, column=2).value = roof["panels"]
        solar.cell(row=row, column=3).value = roof["panel_rating"]
        solar.cell(row=row, column=4).value = array_kw
        solar.cell(row=row, column=5).value = roof["tilt"]
        solar.cell(row=row, column=6).value = roof["azimuth"]

        for col in range(1, 7):
            value(solar.cell(row=row, column=col))

        total_panels += roof["panels"]
        total_kw += array_kw

        row += 1

    solar.cell(row=row, column=1).value = "TOTAL"
    header(solar.cell(row=row, column=1))

    solar.cell(row=row, column=2).value = total_panels
    solar.cell(row=row, column=4).value = total_kw

    value(solar.cell(row=row, column=2))
    value(solar.cell(row=row, column=4))
    
    solar["A1"] = "Solar Model"
    title(solar["A1"])

    solar["A3"] = "Version 1.0 uses estimated monthly solar generation."
    solar["A5"] = "A future version will calculate generation from:"
    solar["A6"] = "• Christchurch irradiation"
    solar["A7"] = "• Roof orientation"
    solar["A8"] = "• Roof pitch"
    solar["A9"] = "• Shading"
    solar["A10"] = "• System losses"

    # --------------------------------------------------
    # HOT WATER
    # --------------------------------------------------

    hotwater["A1"] = "Hot Water"
    title(hotwater["A1"])

    hotwater["A3"] = "Current System"
    header(hotwater["A3"])

    hotwater["A4"] = "Cylinder"
    hotwater["B4"] = "Peter Cocks 300 L"

    hotwater["A5"] = "Elements"
    hotwater["B5"] = "2 × 3 kW"

    hotwater["A6"] = "Thermostat"
    hotwater["B6"] = "70 °C"

    for r in range(4, 7):
        value(hotwater[f"A{r}"])
        value(hotwater[f"B{r}"])

        # --------------------------------------------------
    # BATTERY
    # --------------------------------------------------

    battery["A1"] = "Battery Options"
    title(battery["A1"])

    battery["A3"] = "Battery Size"
    battery["B3"] = "Status"

    header(battery["A3"])
    header(battery["B3"])

    battery_options = [
        ("None Installed", "Current"),
        ("8 kWh", "Future Option"),
        ("16 kWh", "Future Option"),
        ("24 kWh", "Future Option"),
    ]

    row = 4

    for size, status in battery_options:

        battery[f"A{row}"] = size
        battery[f"B{row}"] = status

        value(battery[f"A{row}"])
        value(battery[f"B{row}"])

        row += 1

    # --------------------------------------------------
    # FINANCIAL
    # --------------------------------------------------

    financial["A1"] = "Financial Comparison"
    title(financial["A1"])

    headings = [
        "Option",
        "Capital Cost",
        "Annual Saving",
        "Simple Payback",
    ]

    for col, text in enumerate(headings, start=1):
        cell = financial.cell(row=3, column=col)
        cell.value = text
        header(cell)

    options = [
        "Do Nothing",
        "Shelly Hot Water",
        "New Cylinder",
        "Heat Pump HWC",
        "8 kWh Battery",
        "16 kWh Battery",
        "24 kWh Battery",
    ]

    row = 4

    for option in options:

        financial.cell(row=row, column=1).value = option

        for col in range(1, 5):
            value(financial.cell(row=row, column=col))

        row += 1

    # --------------------------------------------------
    # DASHBOARD
    # --------------------------------------------------

    dashboard["A1"] = "Dashboard"
    title(dashboard["A1"])

    dashboard["A3"] = "System Status"
    header(dashboard["A3"])

    dashboard["A5"] = "Annual Usage"
    dashboard["B5"] = summary["usage"]

    dashboard["A6"] = "Estimated Solar"
    dashboard["B6"] = summary["solar"]

    dashboard["A7"] = "Solar Coverage"
    dashboard["B7"] = summary["coverage"]
    dashboard["B7"].number_format = "0.0%"

    dashboard["A9"] = "Current Recommendation"
    header(dashboard["A9"])

    dashboard["A11"] = "1. Configure Shelly hot water control"
    dashboard["A12"] = "2. Configure EV solar charging"
    dashboard["A13"] = "3. Gather operating data"
    dashboard["A14"] = "4. Review battery after 12 months"

    for r in (5, 6, 7):
        value(dashboard[f"A{r}"])
        value(dashboard[f"B{r}"])

    # --------------------------------------------------
    # SAVE WORKBOOK
    # --------------------------------------------------

    for ws in worksheets:
        autofit_columns(ws)
    
    wb.save(OUTPUT_FILE)

    print(f"KHEO workbook created: {OUTPUT_FILE}")

    return wb


if __name__ == "__main__":
    build_workbook()